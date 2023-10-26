import os
import json
import openpyxl

Script_Path= os.path.dirname(__file__)
Folder_Name = "Json"         #might be changed it's path to folder where are all of the .json files
Folder = os.path.join(Script_Path, Folder_Name)      
lowest_prices = []

def sort_files_by_number(filename):
    parts = filename.split("fragments")     #name of the json files
    if len(parts) > 1:
        return int(parts[1].split(".")[0])
    return 0

pliki = os.listdir(Folder)
pliki.sort(key=sort_files_by_number)

for plik in pliki:
    if plik.endswith(".json"):
        sciezka_pliku = os.path.join(Folder, plik)
        print(sciezka_pliku)
        with open(sciezka_pliku, 'r') as file:
            data = json.load(file)
            if 'lowest_price' in data:
                lowest_price = data['lowest_price']
                lowest_prices.append(lowest_price)

Excel_File = os.path.join(Script_Path, 'steam_data.xlsx') #You can change here excel file where you want to save imported data
workbook = openpyxl.load_workbook(Excel_File)

Sheet = workbook.active

Last_col = Sheet.max_column

New_Column = openpyxl.utils.get_column_letter(Last_col + 1)
First_Row = 0                    #select the starting row for save  can be changed
for i, price in enumerate(lowest_prices):
    Sheet[f'{New_Column}{First_Row + i}'] = price

workbook.save(Excel_File)
